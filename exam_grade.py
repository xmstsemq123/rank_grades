from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

#setting parameter
ws_row_count = 2460 + 1 //名單上總人數
ws_read_col = ['','A','B','C','D','F','H','J','L']

#load excel files
wb = load_workbook('exam_grade.xlsx')
ws = wb['03EE']

#create an new excel file
wb2 = Workbook()
ws2 = wb2.active
ws2.title = '03EE'

#setting ws2 cell style
Color=['c6efce','006100']   #綠
fill_style = PatternFill('solid', fgColor=Color[0])
font_style = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color=Color[1])
Color2 = ['ffeb9c', '9c6500']  #黃
fill_style2 = PatternFill('solid', fgColor=Color2[0])
font_style2 = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color=Color2[1])

#write title in new excel file
ws2.append(['學校名稱','報名序號','類別與群組','國文','英文','數學','專一','專二','加權分數(以台科大電機加權)','全國排名'])
for row in range(1,ws_row_count):
    for col in range(1,len(ws_read_col)):
        char = ws_read_col[col] + str(row)
        char2 = get_column_letter(col) + str(row)
        ws2[char2].value = ws[char].value

#calculate weighted score
for row in range(2,ws_row_count):
    row_str = str(row)
    score = float(ws2['D'+row_str].value)*1+float(ws2['E'+row_str].value)*2+float(ws2['F'+row_str].value)*2+float(ws2['G'+row_str].value)*3+float(ws2['H'+row_str].value)*3
    ws2['I' + row_str].value = score

#rank
for num in range(1,101):
    for row in range(2,ws_row_count):
        score = 'I' + str(row)
        score_next = 'I' + str(row + 1)
        if ws2[score_next].value == None:
            break
        if float(ws2[score_next].value) > float(ws2[score].value):#泡沫排序
            str_before = ''
            str_after = ''
            for col in range(1,len(ws_read_col)+1):
                char = get_column_letter(col) + str(row)
                char2 = get_column_letter(col) + str(row+1)
                str_before = ws2[char].value
                str_after = ws2[char2].value
                ws2[char].value = str_after
                ws2[char2].value = str_before

for num in range(1,101):
    for row in range(2,ws_row_count):
        score = 'I' + str(row)
        score_next = 'I' + str(row + 1)  
        if ws2[score_next].value == None:
            break
        if float(ws2[score_next].value) == float(ws2[score].value):#同分參酌
            str_row = str(row)
            str_row2 = str(row + 1)
            if float(ws2['G' + str_row2].value) > float(ws2['G' + str_row].value):#1專一
                str_before = ''
                str_after = ''
                for col in range(1,len(ws_read_col)+1):
                    char = get_column_letter(col) + str(row)
                    char2 = get_column_letter(col) + str(row+1)
                    str_before = ws2[char].value
                    str_after = ws2[char2].value
                    ws2[char].value = str_after
                    ws2[char2].value = str_before
            elif float(ws2['G' + str_row2].value) < float(ws2['G' + str_row].value):#1專一
                continue
            elif float(ws2['H' + str_row2].value) > float(ws2['H' + str_row].value):#2專二
                str_before = ''
                str_after = ''
                for col in range(1,len(ws_read_col)+1):
                    char = get_column_letter(col) + str(row)
                    char2 = get_column_letter(col) + str(row+1)
                    str_before = ws2[char].value
                    str_after = ws2[char2].value
                    ws2[char].value = str_after
                    ws2[char2].value = str_before
            elif float(ws2['H' + str_row2].value) < float(ws2['H' + str_row].value):#2專二
                continue
            elif float(ws2['E' + str_row2].value) > float(ws2['E' + str_row].value):#3英文
                str_before = ''
                str_after = ''
                for col in range(1,len(ws_read_col)+1):
                    char = get_column_letter(col) + str(row)
                    char2 = get_column_letter(col) + str(row+1)
                    str_before = ws2[char].value
                    str_after = ws2[char2].value
                    ws2[char].value = str_after
                    ws2[char2].value = str_before
            elif float(ws2['E' + str_row2].value) < float(ws2['E' + str_row].value):#3英文
                continue
            elif float(ws2['F' + str_row2].value) > float(ws2['F' + str_row].value):#4數學
                str_before = ''
                str_after = ''
                for col in range(1,len(ws_read_col)+1):
                    char = get_column_letter(col) + str(row)
                    char2 = get_column_letter(col) + str(row+1)
                    str_before = ws2[char].value
                    str_after = ws2[char2].value
                    ws2[char].value = str_after
                    ws2[char2].value = str_before
            elif float(ws2['F' + str_row2].value) < float(ws2['F' + str_row].value):#4數學
                continue
            elif float(ws2['D' + str_row2].value) > float(ws2['D' + str_row].value):#5國文
                str_before = ''
                str_after = ''
                for col in range(1,len(ws_read_col)+1):
                    char = get_column_letter(col) + str(row)
                    char2 = get_column_letter(col) + str(row+1)
                    str_before = ws2[char].value
                    str_after = ws2[char2].value
                    ws2[char].value = str_after
                    ws2[char2].value = str_before

for row in range(2,ws_row_count):
    ws2['J' + str(row)].value = str(row-1)

#cell fillwith color
for row in range(1,ws_row_count):
    if str(ws2['B' + str(row)].value)[0:8] == '5121-308':#電機忠
        for col in range(1,len(ws_read_col)+2):
            char2 = get_column_letter(col) + str(row)
            ws2[char2].fill = fill_style
            ws2[char2].font = font_style
    if str(ws2['B' + str(row)].value)[0:8] == '5121-309':#電機孝
        for col in range(1,len(ws_read_col)+2):
            char2 = get_column_letter(col) + str(row)
            ws2[char2].fill = fill_style2
            ws2[char2].font = font_style2

#save new excel file
wb2.save('111統測電機類排名(台科大電機加權).xlsx')
